"""Read fixed-layout purchase receipt workbooks (B1–B3, B5 header; lines from row 10)."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

_MIN_EXCEL_DATE_SERIAL = 30000


@dataclass
class ReceiptLine:
    row_number: int
    item_code: str
    item_name: str
    quantity: float
    unit_cost: float


@dataclass
class ReceiptWorkbook:
    path: Path
    invoice_date: date
    vendor: str
    branch: str
    vendor_invoice_num: str
    lines: list[ReceiptLine]
    receipt_note: str = ""


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_date_value(value, *, context: str = "") -> date:
    hint = f" ({context})" if context else ""
    if value is None or value == "":
        raise ValueError(f"Invoice date is empty{hint}")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if serial >= _MIN_EXCEL_DATE_SERIAL:
            return from_excel(serial).date()
        raise ValueError(
            f"Invoice date looks like number {value!r}, not a calendar date{hint}"
        )
    s = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return date.fromisoformat(s)
    try:
        serial = float(s)
    except ValueError as exc:
        raise ValueError(f"Unrecognized date: {value!r}{hint}") from exc
    if serial >= _MIN_EXCEL_DATE_SERIAL:
        return from_excel(serial).date()
    raise ValueError(f"Unrecognized date: {value!r}{hint}")


def _cell_float(value, *, context: str) -> float:
    if value is None or value == "":
        raise ValueError(f"{context} is empty")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    s = str(value).strip().replace(",", "")
    try:
        return float(s)
    except ValueError as exc:
        raise ValueError(f"{context} is not a number: {value!r}") from exc


def line_start_row() -> int:
    return int(os.environ.get("ASPIRE_LINE_START_ROW", "10"))


def sheet_name() -> str | None:
    name = os.environ.get("ASPIRE_EXCEL_SHEET", "").strip()
    return name or None


def read_receipt_workbook(path: Path) -> ReceiptWorkbook:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name():
            if sheet_name() not in wb.sheetnames:
                raise ValueError(
                    f"Sheet {sheet_name()!r} not found in {path.name}; "
                    f"available: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name()]
        else:
            ws = wb.active
        if ws is None:
            raise ValueError(f"No worksheet in {path.name}")

        invoice_date = parse_date_value(ws["B1"].value, context="B1")
        vendor = _cell_str(ws["B2"].value)
        branch = _cell_str(ws["B3"].value)
        vendor_invoice_num = _cell_str(ws["B5"].value)
        receipt_note = _cell_str(ws["B6"].value)

        if not vendor:
            raise ValueError("B2 (Vendor) is empty")
        if not branch:
            raise ValueError("B3 (Branch) is empty")
        if not vendor_invoice_num:
            raise ValueError("B5 (Invoice number) is empty")

        start = line_start_row()
        lines: list[ReceiptLine] = []
        row = start
        while True:
            code = _cell_str(ws.cell(row=row, column=2).value)
            name = _cell_str(ws.cell(row=row, column=3).value)
            if not code and not name:
                break
            qty = _cell_float(
                ws.cell(row=row, column=4).value, context=f"D{row} quantity"
            )
            unit_cost = _cell_float(
                ws.cell(row=row, column=5).value, context=f"E{row} unit price"
            )
            if qty <= 0:
                raise ValueError(f"Row {row}: quantity must be > 0 (got {qty})")
            if unit_cost < 0:
                raise ValueError(f"Row {row}: unit price must be >= 0 (got {unit_cost})")
            lines.append(
                ReceiptLine(
                    row_number=row,
                    item_code=code,
                    item_name=name,
                    quantity=qty,
                    unit_cost=unit_cost,
                )
            )
            row += 1

        if not lines:
            raise ValueError(
                f"No line items from row {start} (columns B and C empty)"
            )

        return ReceiptWorkbook(
            path=path,
            invoice_date=invoice_date,
            vendor=vendor,
            branch=branch,
            vendor_invoice_num=vendor_invoice_num,
            lines=lines,
            receipt_note=receipt_note,
        )
    finally:
        wb.close()


def vendor_invoice_datetime_iso(d: date) -> str:
    """Aspire VendorInvoiceDate expects date-time; use noon UTC-neutral local-style."""
    return f"{d.isoformat()}T12:00:00"
