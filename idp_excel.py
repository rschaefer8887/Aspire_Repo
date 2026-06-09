"""Write extracted invoice data to Invoice_Import Excel template."""

from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from idp_costs import (
    LineOutput,
    apply_tax,
    line_total_cost,
    reconcile_line_costs,
)
from idp_openai import ExtractionResult, format_invoice_number
from idp_paths import default_branch, excel_template_path, sanitize_filename_part
from idp_pack_conversion import maybe_convert_box_line, maybe_convert_canister_line
from idp_roll_conversion import maybe_convert_roll_line
from idp_sod import SodSplitResult, build_sod_receipt_note
from idp_vendor_profiles import IDAHO_SOD_PROFILE, VendorProfile, vendor_profile_for


def _ensure_template() -> Path:
    path = excel_template_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file():
        return path
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice_Import"
    ws["A3"], ws["B3"] = "Branch", default_branch()
    wb.save(path)
    wb.close()
    return path


def write_invoice_workbook(
    template_path: Path,
    output_path: Path,
    *,
    invoice_date: date | None,
    vendor_name: str,
    branch: str,
    invoice_number: str,
    lines: list[LineOutput],
    invoice_total: float | None = None,
    receipt_note: str | None = None,
    profile: VendorProfile | None = None,
    sheet_name: str = "Invoice_Import",
    line_start: int = 10,
) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not in template; have: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]

    if invoice_date:
        ws["B1"] = invoice_date
    ws["B2"] = vendor_name
    ws["B3"] = branch
    ws["B5"] = invoice_number
    if receipt_note:
        ws["B6"] = receipt_note

    reconciled = True
    if (
        profile
        and profile.reconcile_to_invoice_total
        and invoice_total is not None
        and invoice_total > 0
    ):
        reconciled = reconcile_line_costs(lines, invoice_total)

    row = line_start
    for line in lines:
        ws.cell(row=row, column=2, value=line.item_code or None)
        ws.cell(row=row, column=3, value=line.item_name or None)
        ws.cell(row=row, column=4, value=line.quantity)
        ws.cell(row=row, column=5, value=line.unit_cost)
        ws.cell(row=row, column=6, value=line_total_cost(line.quantity, line.unit_cost))
        row += 1

    wb.save(output_path)
    wb.close()
    return reconciled


def extraction_to_lines(
    result: ExtractionResult,
    *,
    profile: VendorProfile | None = None,
) -> list[LineOutput]:
    prof = profile or vendor_profile_for(result.vendor_name, result.vendor_raw)
    out: list[LineOutput] = []
    for line in result.lines:
        qty, unit_price, _roll_note = maybe_convert_roll_line(
            line.quantity,
            line.unit_price,
            description_raw=line.description_raw,
            uom_raw=line.uom_raw,
            item_name=line.item_name,
            item_code=line.item_code,
        )
        qty, unit_price, _canister_note = maybe_convert_canister_line(
            qty,
            unit_price,
            description_raw=line.description_raw,
            item_code=line.item_code,
            item_name=line.item_name,
        )
        qty, unit_price, _box_note = maybe_convert_box_line(
            qty,
            unit_price,
            description_raw=line.description_raw,
            item_code=line.item_code,
            item_name=line.item_name,
        )
        out.append(
            LineOutput(
                item_code=line.item_code or "",
                item_name=line.item_name or line.description_raw,
                quantity=qty,
                unit_cost=apply_tax(unit_price, profile=prof),
            )
        )
    return out


def output_filename(vendor_name: str, invoice_number: str) -> str:
    v = sanitize_filename_part(vendor_name, 60)
    inv = sanitize_filename_part(invoice_number, 40)
    return f"{v}_{inv}.xlsx"


def build_output_path(
    output_dir: Path,
    vendor_name: str,
    invoice_number: str,
) -> Path:
    base = output_filename(vendor_name, invoice_number)
    path = output_dir / base
    if not path.exists():
        return path
    stem = path.stem
    n = 2
    while path.exists():
        path = output_dir / f"{stem}_{n}.xlsx"
        n += 1
    return path


def write_from_extraction(
    result: ExtractionResult,
    output_dir: Path,
    *,
    template_path: Path | None = None,
) -> tuple[Path, bool, float]:
    template = template_path or excel_template_path()
    if not template.is_file():
        template = _ensure_template()

    vendor = result.vendor_name or result.vendor_raw or "Unknown_Vendor"
    profile = vendor_profile_for(result.vendor_name, result.vendor_raw)
    inv_num = format_invoice_number(result.invoice_number_raw)
    out_path = build_output_path(output_dir, vendor, inv_num)
    lines = extraction_to_lines(result, profile=profile)
    if not lines:
        raise ValueError("No line items to write")

    receipt_note = result.receipt_note
    if (
        profile.profile_id == IDAHO_SOD_PROFILE.profile_id
        and result.invoice_total
        and isinstance(result.sod_split, SodSplitResult)
    ):
        work = [
            LineOutput(l.item_code, l.item_name, l.quantity, l.unit_cost)
            for l in lines
        ]
        reconcile_line_costs(work, result.invoice_total)
        lines = work
        receipt_note = build_sod_receipt_note(
            result.invoice_total, lines, result.sod_split
        )

    reconciled = write_invoice_workbook(
        template,
        out_path,
        invoice_date=result.invoice_date,
        vendor_name=vendor,
        branch=default_branch(),
        invoice_number=inv_num,
        lines=lines,
        invoice_total=result.invoice_total,
        receipt_note=receipt_note,
        profile=profile,
    )
    col_f_total = round(
        sum(line_total_cost(l.quantity, l.unit_cost) for l in lines), 2
    )
    return out_path, reconciled, col_f_total
