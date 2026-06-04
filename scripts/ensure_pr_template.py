"""Create PR Template.xlsx with Invoice_Import sheet if missing."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from idp_paths import default_branch, excel_template_path, receipts_ready_dir


def ensure_template() -> Path:
    path = excel_template_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file():
        return path
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice_Import"
    ws["A1"], ws["B1"] = "Invoice Date", ""
    ws["A2"], ws["B2"] = "Vendor", ""
    ws["A3"], ws["B3"] = "Branch", default_branch()
    ws["A4"], ws["B4"] = "Inventory Location", "Driggs"
    ws["A5"], ws["B5"] = "Invoice Number", ""
    ws["A9"] = "Line items (row 10+)"
    ws["B9"], ws["C9"], ws["D9"], ws["E9"] = (
        "Item Code",
        "Item Name",
        "Quantity",
        "Unit Cost",
    )
    wb.save(path)
    print(f"Created template: {path}")
    return path


if __name__ == "__main__":
    ensure_template()
