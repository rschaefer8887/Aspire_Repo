"""Unit tests for Excel parsing (no API)."""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from aspire_excel import read_receipt_workbook  # noqa: E402


class TestAspireExcel(unittest.TestCase):
    @patch.dict(os.environ, {"ASPIRE_EXCEL_SHEET": ""}, clear=False)
    def test_read_fixed_layout(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["B1"] = date(2025, 3, 15)
        ws["B2"] = "Acme Supply"
        ws["B3"] = "Main"
        ws["B4"] = "Driggs"
        ws["B5"] = "INV-1001"
        ws["B10"] = "SKU-1"
        ws["D10"] = 2
        ws["E10"] = 10.5
        ws["B11"] = "SKU-2"
        ws["D11"] = 1
        ws["E11"] = 25

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            wb.save(path)
            wb.close()
            doc = read_receipt_workbook(path)
            self.assertEqual(doc.vendor, "Acme Supply")
            self.assertEqual(len(doc.lines), 2)
            self.assertEqual(doc.lines[0].item_code, "SKU-1")
            self.assertEqual(doc.lines[0].quantity, 2.0)
        finally:
            path.unlink(missing_ok=True)

    @patch.dict(os.environ, {"ASPIRE_EXCEL_SHEET": ""}, clear=False)
    def test_read_line_with_name_only(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["B1"] = date(2025, 3, 15)
        ws["B2"] = "Vendor"
        ws["B3"] = "Main"
        ws["B5"] = "INV-2"
        ws["C10"] = "Colorado Spruce"
        ws["D10"] = 1
        ws["E10"] = 50.0
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = Path(tmp.name)
        try:
            wb.save(path)
            wb.close()
            doc = read_receipt_workbook(path)
            self.assertEqual(len(doc.lines), 1)
            self.assertEqual(doc.lines[0].item_code, "")
            self.assertEqual(doc.lines[0].item_name, "Colorado Spruce")
        finally:
            path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
